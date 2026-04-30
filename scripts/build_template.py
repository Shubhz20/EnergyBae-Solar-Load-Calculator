"""
Generate a starter Solar Load template at templates/solar_load_template.xlsx.

This is a reasonable placeholder so the app produces real output even when
the official Energybae template isn't available. Replace the file with the
official one any time - the cell map in utils/excel_handler.py just needs to
match.

Run from the project root:
    python scripts/build_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F6F4A")
SUB_FILL = PatternFill("solid", fgColor="E8F4EE")
INPUT_FILL = PatternFill("solid", fgColor="FFF7CC")
CALC_FILL = PatternFill("solid", fgColor="EAF1FB")
BORDER = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def _label(ws, coord, text, *, bold=False, fill=None, color=None, size=11):
    c = ws[coord]
    c.value = text
    c.font = Font(bold=bold, color=color or "000000", size=size)
    c.alignment = Alignment(horizontal="left", vertical="center")
    if fill is not None:
        c.fill = fill
    c.border = BORDER


def _input(ws, coord):
    c = ws[coord]
    c.fill = INPUT_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER


def _calc(ws, coord, formula, fmt=None):
    c = ws[coord]
    c.value = formula
    c.fill = CALC_FILL
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def build(out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Solar Load"

    # Column widths
    widths = {"A": 2, "B": 32, "C": 28, "D": 22, "E": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header band
    ws.merge_cells("B2:E2")
    _label(ws, "B2", "ENERGYBAE - Solar Load Calculator",
           bold=True, fill=HEADER_FILL, color="FFFFFF", size=16)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:E3")
    _label(ws, "B3", "Customer Details & Load Inputs",
           bold=True, fill=SUB_FILL, size=12)

    # --- Input section -----------------------------------------------------
    inputs = [
        ("B5", "Consumer Name",            "C5"),
        ("B6", "Consumer Number",          "C6"),
        ("B7", "Billing Unit",             "C7"),
        ("B8", "Tariff Category",          "C8"),
        ("B9", "Connected Load (kW)",      "C9"),
        ("B10", "Avg Monthly Consumption (kWh)", "C10"),
    ]
    for label_cell, label_text, input_cell in inputs:
        _label(ws, label_cell, label_text, bold=True)
        _input(ws, input_cell)

    # --- Calculations section (formulas - app must NEVER overwrite these) --
    ws.merge_cells("B12:E12")
    _label(ws, "B12", "System Sizing & Savings (auto-calculated)",
           bold=True, fill=SUB_FILL, size=12)

    # Annual consumption from average monthly
    _label(ws, "B13", "Annual Consumption (kWh)", bold=True)
    _calc(ws, "C13", "=IF(ISNUMBER(C10), C10*12, \"\")", fmt="#,##0")

    # Recommended system size: rule of thumb ~ 120 kWh per kWp per month in MH
    _label(ws, "B14", "Recommended Solar System Size (kWp)", bold=True)
    _calc(ws, "C14", "=IF(ISNUMBER(C10), ROUND(C10/120, 2), \"\")", fmt="0.00")

    # Estimated annual generation (~ 1500 kWh per kWp per year in MH)
    _label(ws, "B15", "Estimated Annual Generation (kWh)", bold=True)
    _calc(ws, "C15", "=IF(ISNUMBER(C14), C14*1500, \"\")", fmt="#,##0")

    # Estimated annual savings at INR 8 per kWh
    _label(ws, "B16", "Estimated Annual Savings (INR)", bold=True)
    _calc(ws, "C16", "=IF(ISNUMBER(C15), C15*8, \"\")", fmt='"INR" #,##0')

    # Estimated system cost at INR 55,000 per kWp
    _label(ws, "B17", "Estimated System Cost (INR)", bold=True)
    _calc(ws, "C17", "=IF(ISNUMBER(C14), C14*55000, \"\")", fmt='"INR" #,##0')

    # Simple payback period
    _label(ws, "B18", "Payback Period (years)", bold=True)
    _calc(ws, "C18", "=IFERROR(C17/C16, \"\")", fmt="0.0")

    # ROI (lifetime 25 yr)
    _label(ws, "B19", "25-Year ROI (%)", bold=True)
    _calc(ws, "C19", "=IFERROR((C16*25 - C17)/C17*100, \"\")", fmt="0.0\"%\"")

    # Footer note
    ws.merge_cells("B21:E21")
    note = ws["B21"]
    note.value = (
        "Yellow cells are inputs (auto-filled from MSEDCL bill). "
        "Blue cells contain formulas - do not edit."
    )
    note.font = Font(italic=True, size=9, color="666666")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    here = Path(__file__).resolve().parent.parent
    build(here / "templates" / "solar_load_template.xlsx")
