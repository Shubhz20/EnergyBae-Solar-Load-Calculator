"""
openpyxl-based writer for the Energybae Solar Load Excel Template.

Design rules:
- We NEVER touch cells that contain formulas.
- We only write to a configurable map of named input cells.
- The workbook is loaded with keep_vba=False and data_only=False so all
  existing formulas, styles, and named ranges remain intact.
"""

import io
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string


# ---------------------------------------------------------------------------
# Default cell map for the Solar Load Template.
#
# These coordinates target the typical "Customer Details / Load Inputs" panel
# at the top of the Energybae Solar Load template. Adjust if the template
# layout changes - this is the ONLY place that needs editing.
# ---------------------------------------------------------------------------
DEFAULT_CELL_MAP: Dict[str, str] = {
    "consumer_name":           "C5",
    "consumer_number":         "C6",
    "billing_unit":            "C7",
    "tariff_category":         "C8",
    "connected_load_kw":       "C9",
    "avg_monthly_consumption": "C10",
}

# Optional: restrict writes to a specific sheet. None -> active sheet.
DEFAULT_SHEET_NAME: Optional[str] = None


def _to_number(value: Any) -> Any:
    """Convert numeric-looking strings to int/float so Excel treats them as numbers."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip().replace(",", "")
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return value  # leave as text


def _is_formula_cell(cell) -> bool:
    v = cell.value
    return isinstance(v, str) and v.startswith("=")


def _split_coord(coord: str) -> Tuple[str, int]:
    col, row = coordinate_from_string(coord)
    return col, row


def detect_template_kind(template_bytes: bytes) -> str:
    """
    Inspect the workbook and return one of:
        "pranay"  - the two-consumer "E-Bill Analysis" layout
        "simple"  - the bundled C5..C10 single-consumer layout
        "unknown" - fall back to "simple" but warn
    """
    wb = load_workbook(io.BytesIO(template_bytes), keep_vba=False, data_only=False)
    ws = wb.active

    def at(coord):
        v = ws[coord].value
        return str(v).strip().lower() if v is not None else ""

    # Pranay template fingerprints
    if (
        "consumer name" in at("B1")
        and "consumer no" in at("B2")
        and "solar pannel used" in at("B7")
        and at("B8") == "sr.no"
        and "units" in at("D8")
        and "month" in at("G8")
    ):
        return "pranay"

    # Simple bundled template fingerprint
    if (
        "consumer name" in at("B5")
        and "consumer no" in at("B6").replace(".", "")
    ):
        return "simple"

    return "unknown"


def fill_template_auto(template_bytes: bytes, data: Dict[str, Any]) -> bytes:
    """
    Smart entry point: detect the template's layout and dispatch to the
    correct writer. Use this from the Streamlit app instead of calling the
    individual fillers.
    """
    kind = detect_template_kind(template_bytes)
    if kind == "pranay":
        # Lazy import to avoid circular module loading at import time.
        from utils.pranay_template import fill_pranay_template

        load = data.get("connected_load_kw")
        if isinstance(load, (int, float)):
            load_str = f"{int(load)}KW" if float(load).is_integer() else f"{load}KW"
        elif load:
            load_str = f"{load}KW"
        else:
            load_str = ""

        right = {
            "consumer_name":   data.get("consumer_name") or "",
            "consumer_no":     data.get("consumer_number") or "",
            "sanct_load":      load_str,
            "connection_type": data.get("tariff_category") or "",
            "monthly_units":   data.get("monthly_units") or [],
            "months":          data.get("months"),
            "current_bill_amount": data.get("current_bill_amount"),
            "fixed_charges":   data.get("fixed_charges", 130),
        }
        # If we only have an average (no per-month history), seed the last
        # month with the average so the AVERAGE() formula still resolves.
        if not right["monthly_units"] and data.get("avg_monthly_consumption"):
            right["monthly_units"] = [data["avg_monthly_consumption"]] * 12
        return fill_pranay_template(template_bytes, right=right)

    # Default / unknown -> simple writer
    return fill_solar_load_template(template_bytes, data)


def fill_solar_load_template(
    template_bytes: bytes,
    data: Dict[str, Any],
    cell_map: Optional[Dict[str, str]] = None,
    sheet_name: Optional[str] = DEFAULT_SHEET_NAME,
) -> bytes:
    """
    Write `data` into the input cells of the Solar Load template.

    Parameters
    ----------
    template_bytes : bytes
        Raw bytes of the .xlsx template.
    data : dict
        Verified field values from the UI.
    cell_map : dict, optional
        Override the default field -> cell mapping.
    sheet_name : str, optional
        Worksheet to write into. Defaults to the active sheet.

    Returns
    -------
    bytes : the filled .xlsx as a byte string ready for download.
    """
    cell_map = cell_map or DEFAULT_CELL_MAP

    wb = load_workbook(filename=io.BytesIO(template_bytes), keep_vba=False)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    numeric_fields = {"connected_load_kw", "avg_monthly_consumption"}

    for field, coord in cell_map.items():
        if field not in data:
            continue
        value = data.get(field)
        if value is None or value == "":
            continue

        cell = ws[coord]

        # Hard rule: never overwrite a formula cell.
        if _is_formula_cell(cell):
            continue

        if field in numeric_fields:
            value = _to_number(value)

        cell.value = value

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Batch / consolidated workbook
# ---------------------------------------------------------------------------

BATCH_COLUMNS: List[Tuple[str, str]] = [
    # (header, field_key_in_row_dict OR formula_token)
    ("Consumer Name",                      "consumer_name"),
    ("Consumer Number",                    "consumer_number"),
    ("Billing Unit",                       "billing_unit"),
    ("Tariff Category",                    "tariff_category"),
    ("Connected Load (kW)",                "connected_load_kw"),
    ("Avg Monthly Consumption (kWh)",      "avg_monthly_consumption"),
    # Computed columns - openpyxl writes the formula, Excel evaluates on open.
    ("Annual Consumption (kWh)",           "FORMULA:annual_consumption"),
    ("Recommended System Size (kWp)",      "FORMULA:system_size"),
    ("Estimated Annual Generation (kWh)",  "FORMULA:annual_generation"),
    ("Estimated Annual Savings (INR)",     "FORMULA:annual_savings"),
    ("Estimated System Cost (INR)",        "FORMULA:system_cost"),
    ("Payback Period (years)",             "FORMULA:payback"),
    ("25-Year ROI (%)",                    "FORMULA:roi_25"),
]

# Solar sizing assumptions - tweak in one place.
GEN_PER_KWP_PER_YEAR = 1500   # kWh
TARIFF_INR_PER_KWH   = 8       # INR
COST_INR_PER_KWP     = 55000   # INR


def _formula_for(token: str, row_idx: int) -> str:
    """Build the cell formula for a computed column on a given Excel row."""
    # Column letters within the row that we reference.
    # Order must match BATCH_COLUMNS above.
    col = {
        "consumer_name":            "A",
        "consumer_number":          "B",
        "billing_unit":             "C",
        "tariff_category":          "D",
        "connected_load_kw":        "E",
        "avg_monthly_consumption":  "F",
        "annual_consumption":       "G",
        "system_size":              "H",
        "annual_generation":        "I",
        "annual_savings":           "J",
        "system_cost":              "K",
    }
    r = row_idx
    if token == "annual_consumption":
        return f'=IF(ISNUMBER({col["avg_monthly_consumption"]}{r}), {col["avg_monthly_consumption"]}{r}*12, "")'
    if token == "system_size":
        return f'=IF(ISNUMBER({col["avg_monthly_consumption"]}{r}), ROUND({col["avg_monthly_consumption"]}{r}/120, 2), "")'
    if token == "annual_generation":
        return f'=IF(ISNUMBER({col["system_size"]}{r}), {col["system_size"]}{r}*{GEN_PER_KWP_PER_YEAR}, "")'
    if token == "annual_savings":
        return f'=IF(ISNUMBER({col["annual_generation"]}{r}), {col["annual_generation"]}{r}*{TARIFF_INR_PER_KWH}, "")'
    if token == "system_cost":
        return f'=IF(ISNUMBER({col["system_size"]}{r}), {col["system_size"]}{r}*{COST_INR_PER_KWP}, "")'
    if token == "payback":
        return f'=IFERROR({col["system_cost"]}{r}/{col["annual_savings"]}{r}, "")'
    if token == "roi_25":
        return f'=IFERROR(({col["annual_savings"]}{r}*25 - {col["system_cost"]}{r}) / {col["system_cost"]}{r} * 100, "")'
    raise ValueError(f"Unknown formula token: {token}")


def _coerce_for_excel(field: str, value: Any) -> Any:
    """Numeric fields -> numbers, others -> strings."""
    if value is None or value == "":
        return None
    if field in {"connected_load_kw", "avg_monthly_consumption"}:
        return _to_number(value)
    return value


def build_batch_workbook(rows: List[Dict[str, Any]]) -> bytes:
    """
    Build a consolidated Excel where each input bill is one row.

    Layout:
        Row 1: Header band + frozen pane.
        Rows 2..N: One row per customer with their extracted values plus
                   per-row formulas for system sizing, savings, payback, ROI.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Solar Load Batch"

    header_fill = PatternFill("solid", fgColor="1F6F4A")
    header_font = Font(bold=True, color="FFFFFF")
    input_fill = PatternFill("solid", fgColor="FFF7CC")
    calc_fill = PatternFill("solid", fgColor="EAF1FB")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header row
    for idx, (header, _key) in enumerate(BATCH_COLUMNS, start=1):
        c = ws.cell(row=1, column=idx, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
    ws.row_dimensions[1].height = 36

    # Data rows
    for ridx, row in enumerate(rows, start=2):
        for cidx, (_header, key) in enumerate(BATCH_COLUMNS, start=1):
            cell = ws.cell(row=ridx, column=cidx)
            if key.startswith("FORMULA:"):
                token = key.split(":", 1)[1]
                cell.value = _formula_for(token, ridx)
                cell.fill = calc_fill
                cell.font = Font(bold=True)
            else:
                cell.value = _coerce_for_excel(key, row.get(key))
                cell.fill = input_fill

    # Number formats per column
    fmt_map = {
        "Connected Load (kW)":              "0.00",
        "Avg Monthly Consumption (kWh)":    "#,##0",
        "Annual Consumption (kWh)":         "#,##0",
        "Recommended System Size (kWp)":    "0.00",
        "Estimated Annual Generation (kWh)":"#,##0",
        "Estimated Annual Savings (INR)":   '"INR" #,##0',
        "Estimated System Cost (INR)":      '"INR" #,##0',
        "Payback Period (years)":           "0.0",
        "25-Year ROI (%)":                  '0.0"%"',
    }
    for cidx, (header, _key) in enumerate(BATCH_COLUMNS, start=1):
        if header in fmt_map:
            for r in range(2, len(rows) + 2):
                ws.cell(row=r, column=cidx).number_format = fmt_map[header]

    # Column widths (approximate; openpyxl can't auto-fit)
    widths = [26, 18, 18, 22, 20, 26, 22, 26, 28, 26, 24, 20, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
