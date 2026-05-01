"""
Map extracted bill data into the Energybae "E-Bill Analysis" template
(structure: Pranay HOME E-Bill Analysis.xlsx).

The template is two consumers wide:
    - Consumer 1 -> labels in B, values in D (units in D, bill in E, cost in F)
    - Consumer 2 -> labels in B, values in H (units in H, bill in I, cost in J)

Header block (rows 1-7):
    D1/H1  Consumer Name
    D2/H2  Consumer No
    D3/H3  Fixed Charges          (default 130)
    D4/H4  Sanct. Load (kW)
    D5/H5  Connection Type
    C7     Solar Panel wattage    (default 600)

Monthly table (rows 8-21):
    Row 8 headers: Sr.No | Month | Units | Bill Amount | Unit Cost (x2)
    Rows 9-20: 12 months of data (Feb -> Jan typically)
    F/J column = formula =(BillAmount - FixedCharges) / Units

Calculations (rows 22-26): averages, recommended kW, solar panel count.
The script ONLY writes input cells; every formula in the template is
preserved.
"""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


# Per-consumer cell map. Add a third side later if the template grows.
CONSUMER_SLOTS = {
    "left":  {  # columns B/C/D/E/F
        "name":          "D1",
        "consumer_no":   "D2",
        "fixed_charges": "D3",
        "sanct_load":    "D4",
        "connection":    "D5",
        "month_col":     "C",
        "units_col":     "D",
        "bill_col":      "E",
        "cost_col":      "F",
        "sr_col":        "B",
    },
    "right": {  # columns G/H/I/J
        "name":          "H1",
        "consumer_no":   "H2",
        "fixed_charges": "H3",
        "sanct_load":    "H4",
        "connection":    "H5",
        "month_col":     "G",
        "units_col":     "H",
        "bill_col":      "I",
        "cost_col":      "J",
        "sr_col":        None,  # right side has no Sr.No column
    },
}

FIRST_DATA_ROW = 9       # row 9 = first month
LAST_DATA_ROW = 20       # row 20 = 12th month (Jan)
DEFAULT_FIXED_CHARGES = 130
DEFAULT_PANEL_WATTAGE = 600


def _to_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(",", "").strip()
    # strip trailing units like "KW", "kWh"
    s = s.split()[0] if s else s
    try:
        return float(s)
    except ValueError:
        # last resort: pull any number out
        import re
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return float(m.group(0)) if m else None


def _write_consumer_block(
    ws,
    slot: str,
    *,
    consumer_name: str,
    consumer_no: str,
    sanct_load: str,
    connection_type: str,
    monthly_units: List[float],
    months: Optional[List[datetime]] = None,
    current_bill_amount: Optional[float] = None,
    fixed_charges: float = DEFAULT_FIXED_CHARGES,
) -> None:
    """Fill one consumer's column block in the template."""
    s = CONSUMER_SLOTS[slot]

    # Header block
    ws[s["name"]] = consumer_name
    ws[s["consumer_no"]] = consumer_no
    ws[s["fixed_charges"]] = fixed_charges
    ws[s["sanct_load"]] = sanct_load
    ws[s["connection"]] = connection_type

    # Monthly table - units only; bill amount only on last row.
    units = list(monthly_units)
    n_rows = LAST_DATA_ROW - FIRST_DATA_ROW + 1   # 12

    # Pad / trim to 12
    if len(units) < n_rows:
        units = [None] * (n_rows - len(units)) + units
    units = units[-n_rows:]

    # Months default: previous Feb -> current Jan
    if months is None:
        # We don't know the year programmatically here without info; let the
        # template's existing month labels stand. Only overwrite if caller
        # passes months explicitly.
        months = [None] * n_rows

    for i, (u, m) in enumerate(zip(units, months)):
        r = FIRST_DATA_ROW + i
        if m is not None:
            ws[f'{s["month_col"]}{r}'] = m
        if u is not None:
            ws[f'{s["units_col"]}{r}'] = float(u)

    if current_bill_amount is not None:
        ws[f'{s["bill_col"]}{LAST_DATA_ROW}'] = float(current_bill_amount)


def fill_pranay_template(
    template_bytes: bytes,
    *,
    left: Optional[Dict[str, Any]] = None,
    right: Optional[Dict[str, Any]] = None,
    panel_wattage: float = DEFAULT_PANEL_WATTAGE,
) -> bytes:
    """
    Fill one or both consumer slots in the Pranay HOME template.

    Each `left` / `right` dict accepts:
        consumer_name, consumer_no, sanct_load, connection_type,
        monthly_units (list of 12 floats, oldest -> newest),
        months (optional list of 12 date-like values),
        current_bill_amount (optional - placed in last month),
        fixed_charges (optional - default 130).
    """
    wb = load_workbook(io.BytesIO(template_bytes), keep_vba=False)
    ws = wb.active

    if panel_wattage:
        ws["C7"] = float(panel_wattage)

    if left:
        _write_consumer_block(ws, "left", **left)
    if right:
        _write_consumer_block(ws, "right", **right)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
