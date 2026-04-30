"""
Smoke tests for utils.excel_handler.

Run with:
    python -m pytest tests/ -q
or just:
    python tests/test_excel_handler.py
"""

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook, load_workbook

from utils.excel_handler import fill_solar_load_template, DEFAULT_CELL_MAP


def _build_dummy_template() -> bytes:
    """Create an in-memory template that mimics the real one:
    - input cells where Gemini values land
    - a formula cell that must NOT be overwritten
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Solar Load"

    ws["B5"], ws["B6"], ws["B7"] = "Consumer Name", "Consumer No.", "Billing Unit"
    ws["B8"], ws["B9"], ws["B10"] = "Tariff", "Connected Load (kW)", "Avg kWh/month"

    # Pre-fill formula cells - these must survive untouched.
    ws["D9"] = "=C9*1.0"        # uses connected load
    ws["D10"] = "=C10*12"       # annualised consumption
    ws["E10"] = "=D10*7"        # rough annual savings

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_fills_inputs_and_preserves_formulas():
    template = _build_dummy_template()
    data = {
        "consumer_name": "ACME Industries",
        "consumer_number": "123456789012",
        "billing_unit": "BU-415",
        "connected_load_kw": "12.5",      # string -> coerced to float
        "tariff_category": "LT-V Industrial",
        "avg_monthly_consumption": "850",
    }

    out_bytes = fill_solar_load_template(template, data)
    wb = load_workbook(io.BytesIO(out_bytes))
    ws = wb.active

    # Inputs written correctly
    assert ws[DEFAULT_CELL_MAP["consumer_name"]].value == "ACME Industries"
    assert ws[DEFAULT_CELL_MAP["consumer_number"]].value == "123456789012"
    assert ws[DEFAULT_CELL_MAP["billing_unit"]].value == "BU-415"
    assert ws[DEFAULT_CELL_MAP["tariff_category"]].value == "LT-V Industrial"
    assert ws[DEFAULT_CELL_MAP["connected_load_kw"]].value == 12.5
    assert ws[DEFAULT_CELL_MAP["avg_monthly_consumption"]].value == 850

    # Formulas preserved
    assert ws["D9"].value == "=C9*1.0"
    assert ws["D10"].value == "=C10*12"
    assert ws["E10"].value == "=D10*7"

    print("OK: inputs written and formulas preserved.")


def test_skips_formula_cells_even_if_mapped():
    """If a cell map ever points at a formula cell, we must NOT overwrite it."""
    template = _build_dummy_template()
    bad_map = {"consumer_name": "D9"}  # D9 holds a formula
    out = fill_solar_load_template(
        template, {"consumer_name": "Should Not Land Here"}, cell_map=bad_map
    )
    ws = load_workbook(io.BytesIO(out)).active
    assert ws["D9"].value == "=C9*1.0"
    print("OK: formula cell protected from accidental overwrite.")


if __name__ == "__main__":
    test_fills_inputs_and_preserves_formulas()
    test_skips_formula_cells_even_if_mapped()
    print("All smoke tests passed.")
